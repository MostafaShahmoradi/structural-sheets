import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import os

class DevelopedPressureTunnelDesign:
    def __init__(self, r_i=1.8, r_s=1.9, r_o=2.1, E_c=20.0, E_s=200.0, E_r=4.0, k_r=1e-6, beta_z=1.0):
        """
        Developed MH coupled design engine based on Schleiss (1997).
        r_i, r_s, r_o: Radii in meters
        E_c, E_s, E_r: Elasticity moduli in GPa
        k_r: Rock permeability in m/s
        beta_z: Concrete tensile strength in MPa
        """
        self.r_i = r_i
        self.r_s = r_s
        self.r_o = r_o
        self.E_c = E_c * 1000.0  # Convert to MPa
        self.E_s = E_s * 1000.0  # Convert to MPa
        self.E_r = E_r * 1000.0  # Convert to MPa
        self.k_r = k_r           # m/s
        self.beta_z = beta_z     # MPa
        
        # Physical constants
        self.nu_c = 0.2
        self.nu_w = 1.01e-6      # Kinematic viscosity of water (m^2/s)
        self.rho_w = 1000.0      # Density of water (kg/m^3)
        self.g = 9.81            # m/s^2

    def calculate_initial_cracking_pressure(self):
        """
        Calculates the internal water pressure (in bar) that triggers initial cracking.
        Based on Eq. 1 & 4 from Schleiss (1997).
        """
        factor = ((2.0 - self.nu_c) / (3.0 * (1.0 - self.nu_c))) * ((1.0 + (self.r_i/self.r_o)**2) / ((self.r_o/self.r_i)**2 - 1.0))
        p_cr_mpa = self.beta_z / factor
        return p_cr_mpa * 10.0  # Convert to bar

    def solve_coupled_state(self, p_i_bar, spacing_cm, bar_diameter_mm, tunnel_above_gw=True, b_or_ak=10.0):
        """
        Performs the full mechanical-hydraulic coupled iterative solver to find:
        - External pressure behind lining p_a (MPa)
        - Seepage discharge q (L/s/m)
        - Crack width 2a (mm)
        - Steel stress sigma_s (MPa)
        - Cracking series factor m
        """
        p_i = p_i_bar * 0.1  # Convert bar to MPa
        p_cr_bar = self.calculate_initial_cracking_pressure()
        
        if p_i_bar < p_cr_bar:
            # Uncracked State
            return {
                "state": "Uncracked",
                "p_a_bar": round(p_i_bar * 0.2, 2),
                "crack_width_mm": 0.0,
                "steel_stress_mpa": 0.0,
                "seepage_l_s_m": 0.0,
                "factor_m": 0.0,
                "cracks_count": 0
            }

        # Determine cracking series and factor m
        # High pressures activate higher cracking series (halving crack spacing d)
        d_orig = spacing_cm / 100.0  # Original spacing in m
        phi = bar_diameter_mm / 1000.0  # Bar diameter in m
        
        # Estimate number of cracking series
        if p_i_bar < 12.0:
            factor_m = 1.0/3.0
            n_cracks = 8
            d_cr = d_orig
        elif p_i_bar < 22.0:
            factor_m = 2.0/3.0
            n_cracks = 16
            d_cr = d_orig / 2.0
        elif p_i_bar < 35.0:
            factor_m = 5.0/6.0
            n_cracks = 32
            d_cr = d_orig / 4.0
        else:
            factor_m = 1.0
            n_cracks = 64
            d_cr = d_orig / 8.0

        # Iterative solver for p_a (using Secant Method)
        # We equilibrate Seepage from Lining (Eq. 5) and Seepage through Rock (Eq. 6 or 7)
        def get_seepage_difference(p_a_test):
            # 1. Calculate steel stress based on compatibility
            A_s = (np.pi * phi**2 / 4.0) / d_orig  # Steel area per meter
            sigma_s = (p_i - p_a_test) * self.r_i / (A_s if A_s > 0 else 1e-4)
            sigma_s = max(0.0, min(sigma_s, 400.0))  # Cap at steel ultimate/yield limit
            
            # 2. Calculate average crack width 2a (Eq. 28)
            crack_width = (sigma_s * 0.8) * d_cr / (3 * (self.E_s)) # in meters
            crack_width_mm = crack_width * 1000.0
            
            # 3. Lining seepage discharge q (Eq. 5) - in m^3/s/m
            head_loss_pa = (p_i - p_a_test) * 1e6 / (self.rho_w * self.g) # Convert MPa to head (m)
            q_lining = (head_loss_pa * (crack_width**3) * n_cracks) / (12.0 * self.nu_w * (self.r_o - self.r_i))
            
            # 4. Rock seepage discharge q (Eq. 6 or 7)
            head_pa = p_a_test * 1e6 / (self.rho_w * self.g)
            if tunnel_above_gw:
                denom = ((1.5 - self.r_o/b_or_ak) / (2.0 * np.pi * self.k_r)) + 1.0
                q_rock = head_pa / (denom if denom != 0 else 1e-4)
            else:
                ln_term = np.log(b_or_ak / self.r_o)
                q_rock = (head_pa - b_or_ak) * 2.0 * np.pi * self.k_r / (ln_term if ln_term > 0 else 1.0)
            
            return q_lining - q_rock, crack_width_mm, sigma_s, q_lining

        # Secant Method loop
        p_a0 = p_i * 0.1
        p_a1 = p_i * 0.5
        
        max_iters = 50
        tol = 1e-5
        p_a = p_a1
        
        for _ in range(max_iters):
            f0, _, _, _ = get_seepage_difference(p_a0)
            f1, w_cr, sig_s, q_l = get_seepage_difference(p_a1)
            
            if abs(f1 - f0) < 1e-12:
                break
                
            p_a_next = p_a1 - f1 * (p_a1 - p_a0) / (f1 - f0)
            p_a_next = max(0.0, min(p_a_next, p_i))  # Bounds
            
            if abs(p_a_next - p_a1) < tol:
                p_a = p_a_next
                break
                
            p_a0, p_a1 = p_a1, p_a_next
            p_a = p_a_next

        # Final check
        _, final_w_cr_mm, final_sig_s, final_q_l = get_seepage_difference(p_a)
        final_q_l_liters = max(0.0, final_q_l * 1000.0) # Convert to L/s/m
        
        return {
            "state": "Cracked",
            "p_a_bar": round(p_a * 10.0, 2),
            "crack_width_mm": round(final_w_cr_mm, 4),
            "steel_stress_mpa": round(final_sig_s, 1),
            "seepage_l_s_m": round(final_q_l_liters, 4),
            "factor_m": round(factor_m, 3),
            "cracks_count": n_cracks
        }

    def export_sensitivity_to_excel(self, file_path, spacing_cm, bar_diameter_mm):
        """
        Runs sensitivity analysis from 0 to 50 bar and exports a professional Excel spreadsheet.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pressure Tunnel Design"
        ws.views.sheetView[0].showGridLines = True
        
        # Styling palettes - Engineering Royal Blue Theme
        navy_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        light_blue = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
        zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        warning_fill = PatternFill(start_color="FDF2F2", end_color="FDF2F2", fill_type="solid")
        
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Segoe UI", size=10)
        font_bold = Font(name="Segoe UI", size=10, bold=True)
        
        border_thin = Side(border_style="thin", color="D1D5DB")
        border_double = Side(border_style="double", color="1B365D")
        border_thick_bottom = Side(border_style="medium", color="1B365D")
        
        grid_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        header_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thick_bottom)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        
        # 1. Project Metadata Block
        ws.merge_cells("A1:G1")
        ws["A1"] = "PRESSURE TUNNEL HYDROMECHANICAL DESIGN SHEET (SCHLEISS 1997)"
        ws["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
        ws["A1"].alignment = align_left
        ws.row_dimensions[1].height = 30
        
        metadata = [\
            ("Lining Inner Radius (ri):", f"{self.r_i} m", "Reinforcement Spacing:", f"{spacing_cm} cm"),\
            ("Lining Outer Radius (ro):", f"{self.r_o} m", "Bar Diameter (phi):", f"{bar_diameter_mm} mm"),\
            ("Concrete Tensile Strength:", f"{self.beta_z} MPa", "Rock Mass Modulus (Er):", f"{self.E_r/1000.0} GPa"),\
            ("Rock Mass Permeability (kr):", f"{self.k_r} m/s", "Steel Elastic Modulus:", f"{self.E_s/1000.0} GPa")\
        ]
        
        for r_idx, row_data in enumerate(metadata, start=3):
            ws.row_dimensions[r_idx].height = 20
            ws.cell(row=r_idx, column=1, value=row_data[0]).font = font_bold
            ws.cell(row=r_idx, column=2, value=row_data[1]).font = font_data
            ws.cell(row=r_idx, column=2).alignment = align_center
            ws.cell(row=r_idx, column=4, value=row_data[2]).font = font_bold
            ws.cell(row=r_idx, column=5, value=row_data[3]).font = font_data
            ws.cell(row=r_idx, column=5).alignment = align_center
            
            for c in range(1, 6):
                ws.cell(row=r_idx, column=c).border = Border(bottom=Side(style="thin", color="E5E7EB"))

        # 2. Table Headers
        headers = [\
            "Internal Pressure\npi (bar)",\
            "External Pressure\npa (bar)",\
            "Lining State\n(Cracked/Uncracked)",\
            "Reinforcement Stress\nsigma_s (MPa)",\
            "Average Crack Width\n2a (mm)",\
            "Seepage Loss Rate\nq (L/s/m)",\
            "Cracks Count\nn"\
        ]
        
        start_row = 9
        ws.row_dimensions[start_row].height = 35
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = navy_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = header_border
            
        # 3. Generate Sensitivity Data & Populate Table
        pressures = range(0, 51, 2)
        for idx, p in enumerate(pressures):
            r = start_row + 1 + idx
            ws.row_dimensions[r].height = 20
            
            res = self.solve_coupled_state(p, spacing_cm, bar_diameter_mm)
            
            ws.cell(row=r, column=1, value=p)
            ws.cell(row=r, column=2, value=float(res["p_a_bar"]))
            ws.cell(row=r, column=3, value=res["state"])
            ws.cell(row=r, column=4, value=float(res["steel_stress_mpa"]))
            ws.cell(row=r, column=5, value=float(res["crack_width_mm"]))
            ws.cell(row=r, column=6, value=float(res["seepage_l_s_m"]))
            ws.cell(row=r, column=7, value=int(res["cracks_count"]))
            
            is_zebra = (idx % 2 == 1)
            row_fill = zebra_fill if is_zebra else PatternFill(fill_type=None)
            
            # Highlight values exceeding allowable limits (Crack width > 0.3mm or Stress > 240 MPa)
            if res["crack_width_mm"] > 0.3 or res["steel_stress_mpa"] > 240.0:
                row_fill = warning_fill
            
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.font = font_data
                cell.border = grid_border
                cell.alignment = align_center
                if row_fill.fill_type:
                    cell.fill = row_fill
                
                if c in [1, 7]:
                    cell.number_format = '#,##0'
                elif c in [2, 4]:
                    cell.number_format = '#,##0.0'
                elif c == 5:
                    cell.number_format = '0.000'
                elif c == 6:
                    cell.number_format = '0.0000'

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row == 1:
                    continue
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        wb.save(file_path)

    def generate_sensitivity_plots(self, image_path, spacing_cm, bar_diameter_mm):
        """
        Generates and saves sensitivity plots for crack width and steel stress.
        """
        pressures = list(range(0, 51, 1))
        crack_widths = []
        steel_stresses = []
        
        for p in pressures:
            res = self.solve_coupled_state(p, spacing_cm, bar_diameter_mm)
            crack_widths.append(res["crack_width_mm"])
            steel_stresses.append(res["steel_stress_mpa"])
            
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
        plt.style.use('seaborn-v0_8-whitegrid')
        
        # Plot 1: Crack Width
        ax1.plot(pressures, crack_widths, color='#2563EB', linewidth=2.5, label=f'D{bar_diameter_mm} @ s{spacing_cm}cm')
        ax1.axhline(y=0.3, color='#EF4444', linestyle='--', linewidth=1.5, label='Allowable Limit (0.3 mm)')
        ax1.set_title("Average Crack Width vs Internal Pressure", fontsize=11, fontweight='bold', color='#1E293B')
        ax1.set_xlabel("Internal Water Pressure (bar)", fontsize=10)
        ax1.set_ylabel("Crack Width (mm)", fontsize=10)
        ax1.set_xlim(0, 50)
        ax1.set_ylim(0, 0.4)
        ax1.legend(loc='upper left', frameon=True)
        ax1.grid(True, linestyle='--', alpha=0.5)
        
        # Plot 2: Steel Stress
        ax2.plot(pressures, steel_stresses, color='#1B365D', linewidth=2.5, label=f'D{bar_diameter_mm} @ s{spacing_cm}cm')
        ax2.axhline(y=240.0, color='#EF4444', linestyle='--', linewidth=1.5, label='Allowable Yield (240 MPa)')
        ax2.set_title("Steel Hoop Stress vs Internal Pressure", fontsize=11, fontweight='bold', color='#1E293B')
        ax2.set_xlabel("Internal Water Pressure (bar)", fontsize=10)
        ax2.set_ylabel("Steel Stress (MPa)", fontsize=10)
        ax2.set_xlim(0, 50)
        ax2.set_ylim(0, 300)
        ax2.legend(loc='upper left', frameon=True)
        ax2.grid(True, linestyle='--', alpha=0.5)
        
        plt.tight_layout()
        plt.savefig(f"{image_path}.png", dpi=150)
        plt.close()

if __name__ == "__main__":
    # Check if we are running in the cloud sandbox environment or a local user PC
    # This prevents FileNotFoundError on Windows computers that don't have /workspace/scratch
    if os.path.exists("/workspace/scratch"):
        excel_out = "/workspace/scratch/Schleiss_Pressure_Tunnel_Design.xlsx"
        plot_out = "/workspace/scratch/Lining_Design_Chart"
    else:
        excel_out = "Schleiss_Pressure_Tunnel_Design.xlsx"
        plot_out = "Lining_Design_Chart"

    engine = DevelopedPressureTunnelDesign()
    
    print("Running Schleiss (1997) Mechanical-Hydraulic Engine...")
    engine.export_sensitivity_to_excel(excel_out, 16, 18)
    print(f"Excel sensitivity report generated: {excel_out}")
    
    engine.generate_sensitivity_plots(plot_out, 16, 18)
    print(f"Design sensitivity charts generated: {plot_out}.png")
    print("Execution Completed Successfully!")
